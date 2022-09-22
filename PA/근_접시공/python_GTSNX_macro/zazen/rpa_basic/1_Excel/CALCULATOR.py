from openpyxl import load_workbook # 파일 불러오기

for x in range(2, 133):
    wb = load_workbook("Volumeloss.xlsx")
    ws = wb.active
    a = ws.cell(row=x, column=3).value
    wb.close()

    wb = load_workbook("calculator.xlsx")
    ws = wb.active
    ws["D2"] = a
    b = ws["H5"].value

    print (b)
    wb.close()

    wb = load_workbook("node_copy.xlsx")
    ws = wb.active
    ws.insert_cols(3, 3 + (x-1)*40 ) # B번째 열로부터 3열 추가
    wb.save("node_copy.xlsx")



# # cell 갯수를 모를 때
# for x in range(1, ws.max_row + 1):
#     for y in range(1, ws.max_column + 1):
#         print(ws.cell(row=x, column=y).value, end=" ") # 1 2 3 4 ..
#     print()

wb.save("sample.xlsx")
wb.close()