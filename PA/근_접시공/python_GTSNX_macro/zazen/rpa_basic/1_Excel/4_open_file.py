from openpyxl import load_workbook # 파일 불러오기
wb = load_workbook("Volumeloss.xlsx") # sample.xlsx 파일에서 wb 을 불러옴
ws = wb.active # 활성화된 Sheet

# cell 데이터 불러오기
# for x in range(1, 11):
#     for y in range(1, 11):
#         print(ws.cell(row=x, column=y).value, end=" ") # 1 2 3 4 ..
#     print()

for x in range(2, 133):
    a = ws.cell(row=x, column=3).value # 1 2 3 4 ..
    print(a)

# # cell 갯수를 모를 때
# for x in range(1, ws.max_row + 1):
#     for y in range(1, ws.max_column + 1):
#         print(ws.cell(row=x, column=y).value, end=" ") # 1 2 3 4 ..
#     print()