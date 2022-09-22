import numpy as np
import math
import pandas as pd
from openpyxl import load_workbook # 파일 불러오기

# Volume Loss cell open
wb = load_workbook("Volume_Loss.xlsx") # sample.xlsx 파일에서 wb 을 불러옴
ws = wb.active # 활성화된 Sheet

diameter = 10
radius = diameter / 2

# First value volume loss
x = 3
Volume_loss = -ws.cell(row=x, column=3).value  # 1 2 3 4 ..
af_radius = math.sqrt(1 - (Volume_loss / 100)) * radius


# first term of xy from one volume_loss
i = 1
bef_contraction_x = radius * math.sin(2 * math.pi / 40 * (i - 1))
bef_contraction_y = radius * math.cos(2 * math.pi / 40 * (i - 1)) + radius
af_contraction_x = af_radius * math.sin(2 * math.pi / 40 * (i - 1))
af_contraction_y = af_radius * math.cos(2 * math.pi / 40 * (i - 1)) + af_radius
Tx = af_contraction_x - bef_contraction_x
Ty = af_contraction_y - bef_contraction_y
T1 = [[Tx, Ty]]

for i in range(2, 41):
    bef_contraction_x = radius * math.sin(2 * math.pi / 40 * (i - 1))
    bef_contraction_y = radius * math.cos(2 * math.pi / 40 * (i - 1)) + radius
    af_contraction_x = af_radius * math.sin(2 * math.pi / 40 * (i - 1))
    af_contraction_y = af_radius * math.cos(2 * math.pi / 40 * (i - 1)) + af_radius
    Tx = af_contraction_x - bef_contraction_x
    Ty = af_contraction_y - bef_contraction_y
    T2 = np.array([[Tx, Ty]])
    T1 = np.concatenate([T1, T2], axis=0)

T0 = np.transpose(T1)

# rest of volume loss value from xl
for x in range(4, 204):
    Volume_loss = -ws.cell(row=x, column=3).value # 1 2 3 4 ..
    af_radius = math.sqrt(1-(Volume_loss/100))*radius

    # 처음 1번
    i = 1
    bef_contraction_x = radius*math.sin(2 * math.pi / 40 * (i-1))
    bef_contraction_y = radius*math.cos(2 * math.pi / 40 * (i-1)) + radius
    af_contraction_x = af_radius*math.sin(2 * math.pi / 40 * (i-1))
    af_contraction_y = af_radius*math.cos(2 * math.pi / 40 * (i-1)) + af_radius
    Tx = af_contraction_x - bef_contraction_x
    Ty = af_contraction_y - bef_contraction_y
    T1 = [[Tx, Ty]]

    for i in range(2, 41):
        bef_contraction_x = radius*math.sin(2 * math.pi / 40 * (i-1))
        bef_contraction_y = radius*math.cos(2 * math.pi / 40 * (i-1)) + radius
        af_contraction_x = af_radius*math.sin(2 * math.pi / 40 * (i-1))
        af_contraction_y = af_radius*math.cos(2 * math.pi / 40 * (i-1)) + af_radius
        Tx = af_contraction_x - bef_contraction_x
        Ty = af_contraction_y - bef_contraction_y
        T2 = np.array([[Tx, Ty]])
        T1 = np.concatenate([T1, T2], axis=0)
    T00 = np.transpose(T1)
    T0 = np.concatenate([T0, T00], axis=0)
df = pd.DataFrame(T0).T
# df.to_excel(excel_writer="/Users/jaeeuncho/Library/CloudStorage/OneDrive-한양대학교/조재은/학교수업 석사3기/paper/기존터널_근접시공/python_GTSNX_macro/zazen/VL2Coord_reverse.xlsx")
df.to_excel(excel_writer="Volumeloss2xycoord.xlsx")
